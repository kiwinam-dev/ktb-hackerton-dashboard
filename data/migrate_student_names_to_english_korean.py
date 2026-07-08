#!/usr/bin/env python3
"""
Firestore students 컬렉션의 기존 학생 이름을
'국문명' -> '영문명(국문명)' 형식으로 안전하게 갱신합니다.

중요:
- 기존 Firestore 문서 ID와 id 필드는 변경하지 않습니다.
- name 필드만 변경합니다.
- 기본값은 DRY RUN이며, --commit 옵션을 줄 때만 실제로 갱신합니다.
- Excel 기준 열: B=과정명, C=국문명, D=영문명, H=생년월일

실행 예시:
1) 변경 대상만 확인
python migrate_student_names_to_english_korean.py \
  --excel "제목 없는 스프레드시트.xlsx" \
  --service-account "./firebase-service-account.json"

2) 실제 변경
python migrate_student_names_to_english_korean.py \
  --excel "제목 없는 스프레드시트.xlsx" \
  --service-account "./firebase-service-account.json" \
  --commit

3) 시트명에서 기수를 읽을 수 없을 때
python migrate_student_names_to_english_korean.py \
  --excel "제목 없는 스프레드시트.xlsx" \
  --service-account "./firebase-service-account.json" \
  --generation 3 \
  --commit
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional
# pyrefly: ignore [missing-import]
import firebase_admin
# pyrefly: ignore [missing-import]
from firebase_admin import credentials, firestore
from openpyxl import load_workbook


COL_COURSE = 2       # B: 과정명
COL_KOREAN_NAME = 3  # C: 국문명
COL_ENGLISH_NAME = 4 # D: 영문명
COL_BIRTHDATE = 8    # H: 생년월일

COLLECTION_NAME = "students"
BATCH_LIMIT = 500


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Firestore students 컬렉션의 name 필드를 영문명(국문명) 형식으로 갱신합니다."
    )
    parser.add_argument("--excel", required=True, help="입력 Excel 파일 경로(.xlsx)")
    parser.add_argument(
        "--service-account",
        required=True,
        help="Firebase Admin SDK 서비스 계정 JSON 파일 경로",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="대상 시트 이름. 생략하면 첫 번째 시트를 사용합니다.",
    )
    parser.add_argument(
        "--generation",
        default=None,
        help="기수. 생략하면 시트명에서 숫자를 추출합니다. 예: P-3기 -> 3",
    )
    parser.add_argument(
        "--course",
        default=None,
        help="과정명을 강제로 지정합니다. 생략하면 B열 과정명을 사용합니다.",
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="실제 Firestore name 필드를 갱신합니다. 생략하면 미리보기만 실행합니다.",
    )
    parser.add_argument(
        "--report-csv",
        default="student_name_migration_report.csv",
        help="변경/미발견/오류 결과를 저장할 CSV 파일 경로",
    )
    return parser.parse_args()


def as_clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def format_display_name(english_name: str, korean_name: str) -> str:
    return f"{english_name}({korean_name})"


def parse_generation(sheet_name: str) -> str:
    match = re.search(r"(\d+)\s*기", sheet_name)
    if not match:
        raise ValueError(
            f"시트명 '{sheet_name}'에서 기수를 추출하지 못했습니다. "
            "예: '--generation 3' 옵션으로 직접 지정하세요."
        )
    return match.group(1)


def normalize_birthdate(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%y%m%d")
    if isinstance(value, date):
        return value.strftime("%y%m%d")

    raw = as_clean_text(value)
    if not raw:
        raise ValueError("생년월일이 비어 있습니다.")

    digits = re.sub(r"\D", "", raw)
    if len(digits) == 8:
        try:
            parsed = datetime.strptime(digits, "%Y%m%d")
        except ValueError as exc:
            raise ValueError(f"유효하지 않은 생년월일입니다: {raw}") from exc
        return parsed.strftime("%y%m%d")

    if len(digits) == 6:
        try:
            datetime.strptime(digits, "%y%m%d")
        except ValueError as exc:
            raise ValueError(f"유효하지 않은 생년월일입니다: {raw}") from exc
        return digits

    raise ValueError(f"생년월일 형식을 해석하지 못했습니다: '{raw}'")


def build_legacy_document_id(generation: int, course: str, korean_name: str, birthdate: str) -> str:
    """기존 import 스크립트의 문서 ID 규칙과 동일하게 생성합니다."""
    return f"{generation}_{course}_{korean_name}_{birthdate}".replace("/", "_").strip()


def read_name_migration_rows(
    excel_path: Path,
    sheet_name: Optional[str],
    forced_generation: Optional[str],
    forced_course: Optional[str],
) -> tuple[list[dict[str, str]], list[dict[str, str]], str, int]:
    workbook = load_workbook(excel_path, data_only=True, read_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"시트 '{sheet_name}'를 찾을 수 없습니다. "
                f"사용 가능한 시트: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    generation = int(forced_generation or parse_generation(worksheet.title))
    rows: list[dict[str, str]] = []
    errors: list[dict[str, str]] = []
    seen_document_ids: set[str] = set()

    for row_number, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        course_from_excel = as_clean_text(row[COL_COURSE - 1] if len(row) >= COL_COURSE else None)
        korean_name = as_clean_text(row[COL_KOREAN_NAME - 1] if len(row) >= COL_KOREAN_NAME else None)
        english_name = as_clean_text(row[COL_ENGLISH_NAME - 1] if len(row) >= COL_ENGLISH_NAME else None)
        birthdate_value = row[COL_BIRTHDATE - 1] if len(row) >= COL_BIRTHDATE else None

        if not any([course_from_excel, korean_name, english_name, as_clean_text(birthdate_value)]):
            continue

        try:
            course = forced_course.strip() if forced_course else course_from_excel
            if not course:
                raise ValueError("B열 과정명이 비어 있습니다.")
            if not korean_name:
                raise ValueError("C열 국문명이 비어 있습니다.")
            if not english_name:
                raise ValueError("D열 영문명이 비어 있습니다.")

            birthdate = normalize_birthdate(birthdate_value)
            document_id = build_legacy_document_id(generation, course, korean_name, birthdate)
            if document_id in seen_document_ids:
                raise ValueError(f"중복 문서 ID입니다: {document_id}")
            seen_document_ids.add(document_id)

            rows.append(
                {
                    "row": str(row_number),
                    "document_id": document_id,
                    "course": course,
                    "korean_name": korean_name,
                    "english_name": english_name,
                    "birthdate": birthdate,
                    "new_name": format_display_name(english_name, korean_name),
                }
            )
        except ValueError as exc:
            errors.append(
                {
                    "row": str(row_number),
                    "document_id": "",
                    "course": course_from_excel,
                    "korean_name": korean_name,
                    "english_name": english_name,
                    "birthdate": as_clean_text(birthdate_value),
                    "new_name": "",
                    "status": "ERROR",
                    "message": str(exc),
                }
            )

    workbook.close()
    return rows, errors, worksheet.title, generation


def initialize_firestore(service_account_path: Path):
    if not firebase_admin._apps:
        credential = credentials.Certificate(str(service_account_path))
        firebase_admin.initialize_app(credential)
    return firestore.client()


def inspect_documents(db, rows: list[dict[str, str]]) -> tuple[list[dict[str, str]], list[dict[str, str]]]:
    """
    기존 문서를 문서 ID로 조회합니다.

    안전 장치:
    - Firestore의 현재 name이 국문명 또는 이미 원하는 표시명일 때만 갱신 대상으로 잡습니다.
    - 다른 값이면 SKIPPED_MISMATCH로 보고하고 자동 변경하지 않습니다.
    """
    updates: list[dict[str, str]] = []
    report: list[dict[str, str]] = []

    for row in rows:
        snapshot = db.collection(COLLECTION_NAME).document(row["document_id"]).get()
        if not snapshot.exists:
            report.append({**row, "old_name": "", "status": "NOT_FOUND", "message": "문서 ID가 Firestore에 없습니다."})
            continue

        data = snapshot.to_dict() or {}
        current_name = as_clean_text(data.get("name"))
        expected_old_name = row["korean_name"]
        new_name = row["new_name"]

        if current_name == new_name:
            report.append({**row, "old_name": current_name, "status": "ALREADY_UPDATED", "message": "이미 원하는 형식입니다."})
            continue

        if current_name != expected_old_name:
            report.append(
                {
                    **row,
                    "old_name": current_name,
                    "status": "SKIPPED_MISMATCH",
                    "message": "Firestore의 현재 name이 엑셀 국문명과 다릅니다. 자동 변경하지 않았습니다.",
                }
            )
            continue

        updates.append({**row, "old_name": current_name})
        report.append({**row, "old_name": current_name, "status": "READY", "message": "name 필드 갱신 대상입니다."})

    return updates, report


def apply_name_updates(db, updates: list[dict[str, str]]) -> int:
    committed_count = 0
    for start in range(0, len(updates), BATCH_LIMIT):
        chunk = updates[start : start + BATCH_LIMIT]
        batch = db.batch()
        for update in chunk:
            document_ref = db.collection(COLLECTION_NAME).document(update["document_id"])
            batch.update(document_ref, {"name": update["new_name"]})
        batch.commit()
        committed_count += len(chunk)
        print(f"[완료] {committed_count}/{len(updates)}건 name 필드 갱신 완료")
    return committed_count


def save_report_csv(report: list[dict[str, str]], output_path: Path) -> None:
    fieldnames = [
        "row", "document_id", "course", "korean_name", "english_name", "birthdate",
        "old_name", "new_name", "status", "message",
    ]
    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(report)


def main() -> int:
    args = parse_args()
    excel_path = Path(args.excel).expanduser().resolve()
    service_account_path = Path(args.service_account).expanduser().resolve()
    report_csv_path = Path(args.report_csv).expanduser().resolve()

    if not excel_path.is_file():
        print(f"[오류] Excel 파일을 찾을 수 없습니다: {excel_path}", file=sys.stderr)
        return 1
    if not service_account_path.is_file():
        print(f"[오류] 서비스 계정 JSON 파일을 찾을 수 없습니다: {service_account_path}", file=sys.stderr)
        return 1

    try:
        rows, input_errors, sheet_name, generation = read_name_migration_rows(
            excel_path, args.sheet, args.generation, args.course
        )
        db = initialize_firestore(service_account_path)
        updates, report = inspect_documents(db, rows)
    except Exception as exc:
        print(f"[오류] 사전 검증 실패: {exc}", file=sys.stderr)
        return 1

    report.extend(input_errors)
    save_report_csv(report, report_csv_path)

    status_counts: dict[str, int] = {}
    for item in report:
        status = item["status"]
        status_counts[status] = status_counts.get(status, 0) + 1

    print() 
    print("===== 이름 마이그레이션 미리보기 =====")
    print(f"시트: {sheet_name}")
    print(f"기수: {generation}")
    print(f"Excel 정상 행: {len(rows)}건")
    for status, count in sorted(status_counts.items()):
        print(f"{status}: {count}건")
    print(f"리포트 파일: {report_csv_path}")

    for item in updates[:5]:
        print(f"- students/{item['document_id']}: {item['old_name']} -> {item['new_name']}")
    if len(updates) > 5:
        print(f"... 외 {len(updates) - 5}건")

    if not args.commit:
        print() 
        print("[DRY RUN] Firestore에는 아직 변경 사항을 저장하지 않았습니다.")
        print("실제 반영 전 CSV 리포트에서 NOT_FOUND / SKIPPED_MISMATCH를 확인하세요.")
        print("실제 변경하려면 명령 끝에 --commit 옵션을 추가하세요.")
        return 0

    if not updates:
        print("[종료] 실제로 갱신할 READY 대상이 없습니다.")
        return 0

    try:
        committed_count = apply_name_updates(db, updates)
    except Exception as exc:
        print(f"[오류] Firestore 갱신 실패: {exc}", file=sys.stderr)
        return 1

    print() 
    print(f"[성공] students 컬렉션의 name 필드 {committed_count}건을 갱신했습니다.")
    print("문서 ID와 id 필드는 변경하지 않았습니다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
