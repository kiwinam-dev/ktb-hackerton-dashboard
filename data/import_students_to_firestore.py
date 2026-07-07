#!/usr/bin/env python3
"""
Excel 학생 명단을 Cloud Firestore의 students 컬렉션에 일괄 등록합니다.

입력 엑셀 기준 열:
- A열: ID(순번)
- B열: 과정명
- C열: 국문명
- H열: 생년월일

문서 ID 형식:
{generation}_{course}_{name}_{YYMMDD}
예: 3_풀스택_가한솔_990824

실행 예시:
1) 미리보기만 실행(기본값, Firestore에 쓰지 않음)
python import_students_to_firestore.py \
  --excel "제목 없는 스프레드시트.xlsx" \
  --service-account "./service-account.json"

2) 실제 Firestore 등록
python import_students_to_firestore.py \
  --excel "제목 없는 스프레드시트.xlsx" \
  --service-account "./service-account.json" \
  --commit

3) 과정명을 모두 "인공지능"으로 강제
python import_students_to_firestore.py \
  --excel "제목 없는 스프레드시트.xlsx" \
  --service-account "./service-account.json" \
  --course "인공지능" \
  --commit
"""

from __future__ import annotations

import argparse
import csv
import re
import sys
from datetime import date, datetime
from pathlib import Path
from typing import Any, Optional

# pyrefly: ignore [missing-import]
import firebase_admin
# pyrefly: ignore [missing-import]
from firebase_admin import credentials, firestore
from openpyxl import load_workbook


COL_ID = 1          # A: 순번
COL_COURSE = 2      # B: 과정명
COL_NAME = 3        # C: 국문명
COL_BIRTHDATE = 8   # H: 생년월일

COLLECTION_NAME = "students"
BATCH_LIMIT = 500  # Firestore WriteBatch의 최대 작업 수


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Excel 학생 명단을 Firestore students 컬렉션에 일괄 등록합니다."
    )
    parser.add_argument(
        "--excel",
        required=True,
        help="입력 Excel 파일 경로(.xlsx)",
    )
    parser.add_argument(
        "--service-account",
        required=True,
        help="Firebase Admin SDK 서비스 계정 JSON 파일 경로",
    )
    parser.add_argument(
        "--sheet",
        default=None,
        help="대상 시트 이름. 생략하면 첫 번째 시트를 사용합니다.",
    )
    parser.add_argument(
        "--generation",
        default=None,
        help="기수. 생략하면 시트명에서 숫자를 추출합니다. 예: P-3기 -> 3",
    )
    parser.add_argument(
        "--course",
        default=None,
        help="과정명을 강제로 지정합니다. 예: 인공지능. 생략하면 B열 과정명을 사용합니다.",
    )
    parser.add_argument(
        "--commit",
        action="store_true",
        help="실제 Firestore에 저장합니다. 생략하면 미리보기만 실행합니다.",
    )
    parser.add_argument(
        "--errors-csv",
        default="students_import_errors.csv",
        help="검증 실패 행을 저장할 CSV 파일 경로",
    )
    return parser.parse_args()


def as_clean_text(value: Any) -> str:
    """Excel 셀 값을 공백 제거한 문자열로 변환합니다."""
    if value is None:
        return ""
    return str(value).strip()


def parse_generation(sheet_name: str) -> str:
    """
    시트명에서 기수를 추출합니다.
    예: P-3기, 3기, AI_12기 -> 3, 3, 12
    """
    match = re.search(r"(\d+)\s*기", sheet_name)
    if not match:
        raise ValueError(
            f"시트명 '{sheet_name}'에서 기수를 추출하지 못했습니다. "
            "예: '--generation 3' 옵션으로 직접 지정하세요."
        )
    return match.group(1)


def normalize_birthdate(value: Any) -> str:
    """
    생년월일을 YYMMDD 형식으로 정규화합니다.

    허용 예:
    - 1999.08.24
    - 1999-08-24
    - 19990824
    - 990824
    - Excel 날짜 셀
    """
    if isinstance(value, datetime):
        return value.strftime("%y%m%d")
    if isinstance(value, date):
        return value.strftime("%y%m%d")

    raw = as_clean_text(value)
    if not raw:
        raise ValueError("생년월일이 비어 있습니다.")

    digits = re.sub(r"\D", "", raw)

    if len(digits) == 8:
        try:
            parsed = datetime.strptime(digits, "%Y%m%d")
        except ValueError as exc:
            raise ValueError(f"유효하지 않은 생년월일입니다: {raw}") from exc
        return parsed.strftime("%y%m%d")

    if len(digits) == 6:
        # 이미 YYMMDD로 입력된 경우로 처리합니다.
        try:
            datetime.strptime(digits, "%y%m%d")
        except ValueError as exc:
            raise ValueError(f"유효하지 않은 생년월일입니다: {raw}") from exc
        return digits

    raise ValueError(
        f"생년월일 형식을 해석하지 못했습니다: '{raw}'. "
        "예: 1999.08.24 또는 990824 형식이어야 합니다."
    )


def build_document_id(generation: str, course: str, name: str, birthdate: str) -> str:
    """
    Firestore 문서 ID를 생성합니다.
    Firestore 문서 ID에 사용할 수 없는 '/' 문자는 '_'로 치환합니다.
    """
    raw = f"{generation}_{course}_{name}_{birthdate}"
    return raw.replace("/", "_").strip()


def read_students(
    excel_path: Path,
    sheet_name: Optional[str],
    forced_generation: Optional[str],
    forced_course: Optional[str],
) -> tuple[list[dict[str, Any]], list[dict[str, Any]], str, str]:
    """Excel을 읽어 Firestore 문서 목록과 오류 목록을 반환합니다."""
    workbook = load_workbook(excel_path, data_only=True, read_only=True)

    if sheet_name:
        if sheet_name not in workbook.sheetnames:
            raise ValueError(
                f"시트 '{sheet_name}'를 찾을 수 없습니다. "
                f"사용 가능한 시트: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook[workbook.sheetnames[0]]

    generation = int(forced_generation or parse_generation(worksheet.title))
    students: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen_document_ids: set[str] = set()

    for row_number, row in enumerate(
        worksheet.iter_rows(min_row=2, values_only=True),
        start=2,
    ):
        excel_id = as_clean_text(row[COL_ID - 1] if len(row) >= COL_ID else None)
        source_course = as_clean_text(
            row[COL_COURSE - 1] if len(row) >= COL_COURSE else None
        )
        name = as_clean_text(row[COL_NAME - 1] if len(row) >= COL_NAME else None)
        birthdate_value = row[COL_BIRTHDATE - 1] if len(row) >= COL_BIRTHDATE else None

        # 완전히 비어 있는 행은 무시합니다.
        if not any([excel_id, source_course, name, as_clean_text(birthdate_value)]):
            continue

        try:
            if not excel_id:
                raise ValueError("A열 ID가 비어 있습니다.")
            if not name:
                raise ValueError("C열 국문명이 비어 있습니다.")

            course = forced_course.strip() if forced_course else source_course
            if not course:
                raise ValueError("B열 과정명이 비어 있습니다.")

            birthdate = normalize_birthdate(birthdate_value)
            document_id = build_document_id(generation, course, name, birthdate)

            if document_id in seen_document_ids:
                raise ValueError(
                    f"중복 문서 ID입니다: {document_id}. "
                    "같은 기수·과정·이름·생년월일 조합이 여러 행에 있습니다."
                )
            seen_document_ids.add(document_id)

            students.append(
                {
                    "document_id": document_id,
                    "data": {
                        "birthdate": birthdate,
                        "course": course,
                        "generation": generation,
                        "id": document_id,
                        "isAdmin": False,
                        "name": name,
                    },
                }
            )
        except ValueError as exc:
            errors.append(
                {
                    "row": row_number,
                    "id": excel_id,
                    "course": source_course,
                    "name": name,
                    "birthdate": as_clean_text(birthdate_value),
                    "reason": str(exc),
                }
            )

    workbook.close()
    return students, errors, worksheet.title, generation


def save_errors_csv(errors: list[dict[str, Any]], output_path: Path) -> None:
    if not errors:
        return

    with output_path.open("w", encoding="utf-8-sig", newline="") as file:
        writer = csv.DictWriter(
            file,
            fieldnames=["row", "id", "course", "name", "birthdate", "reason"],
        )
        writer.writeheader()
        writer.writerows(errors)


def initialize_firestore(service_account_path: Path):
    if not firebase_admin._apps:
        credential = credentials.Certificate(str(service_account_path))
        firebase_admin.initialize_app(credential)
    return firestore.client()


def write_students(db, students: list[dict[str, Any]]) -> int:
    """
    Firestore에 최대 500건 단위로 batch write 합니다.
    동일한 document_id가 이미 존재하면 해당 문서를 동일 데이터로 덮어씁니다.
    """
    committed_count = 0

    for start in range(0, len(students), BATCH_LIMIT):
        chunk = students[start : start + BATCH_LIMIT]
        batch = db.batch()

        for student in chunk:
            document_ref = db.collection(COLLECTION_NAME).document(student["document_id"])
            batch.set(document_ref, student["data"], merge=False)

        batch.commit()
        committed_count += len(chunk)
        print(f"[완료] {committed_count}/{len(students)}건 Firestore 저장 완료")

    return committed_count


def main() -> int:
    args = parse_args()

    excel_path = Path(args.excel).expanduser().resolve()
    service_account_path = Path(args.service_account).expanduser().resolve()
    errors_csv_path = Path(args.errors_csv).expanduser().resolve()

    if not excel_path.is_file():
        print(f"[오류] Excel 파일을 찾을 수 없습니다: {excel_path}", file=sys.stderr)
        return 1

    if not service_account_path.is_file():
        print(
            f"[오류] 서비스 계정 JSON 파일을 찾을 수 없습니다: {service_account_path}",
            file=sys.stderr,
        )
        return 1

    try:
        students, errors, sheet_name, generation = read_students(
            excel_path=excel_path,
            sheet_name=args.sheet,
            forced_generation=args.generation,
            forced_course=args.course,
        )
    except Exception as exc:
        print(f"[오류] Excel 읽기 실패: {exc}", file=sys.stderr)
        return 1

    if errors:
        save_errors_csv(errors, errors_csv_path)
        print(f"[주의] 검증 실패 {len(errors)}건: {errors_csv_path}")

    print()
    print("===== 가져오기 미리보기 =====")
    print(f"시트: {sheet_name}")
    print(f"기수: {generation}")
    print(f"정상 데이터: {len(students)}건")
    print(f"검증 실패: {len(errors)}건")

    for student in students[:5]:
        print(f"- students/{student['document_id']}")
        print(f"  {student['data']}")

    if len(students) > 5:
        print(f"... 외 {len(students) - 5}건")

    if not students:
        print("[종료] 저장할 정상 데이터가 없습니다.", file=sys.stderr)
        return 1

    if not args.commit:
        print()
        print("[DRY RUN] Firestore에는 아직 저장하지 않았습니다.")
        print("실제 저장하려면 명령 끝에 --commit 옵션을 추가하세요.")
        return 0

    try:
        db = initialize_firestore(service_account_path)
        committed_count = write_students(db, students)
    except Exception as exc:
        print(f"[오류] Firestore 저장 실패: {exc}", file=sys.stderr)
        return 1

    print()
    print(f"[성공] students 컬렉션에 {committed_count}건을 저장했습니다.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
